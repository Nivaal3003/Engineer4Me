"""Deterministic and formula-safe Step 110 datasheet exports.

The renderer accepts only a fully validated :class:`DatasheetRevisionRecord`.
It produces the complete canonical JSON record and a presentation workbook
whose cells never contain executable formulas.  Export bytes stay in memory;
this module does not accept paths, write files, approve designs, or claim
standards conformity.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import UTC
from hashlib import sha256
from io import BytesIO
import json
import re
from typing import Literal, Self
from uuid import UUID
from zipfile import ZIP_STORED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import Field, model_validator

from app.engineering.calculations.models import (
    CalculationModel,
    EngineeringQuantity,
    FingerprintText,
    Identifier,
    VersionText,
)
from app.engineering.calculations.units import format_quantity_value
from app.engineering.design.datasheet_models import (
    DatasheetFieldState,
    DatasheetRevisionRecord,
)
from app.engineering.design.datasheet_registry import (
    DEFAULT_DATASHEET_TEMPLATE_REGISTRY,
    PRESSURE_RELIEF_TEMPLATE_ID,
    DatasheetTemplateRegistryError,
)


DATASHEET_EXPORT_VERSION = "1.0.0"
DATASHEET_EXPORT_SCHEMA = "engineer4me.datasheet.export.v1"
DATASHEET_EXPORT_CANONICALIZATION = (
    "engineer4me.datasheet.export.canonical-json.sha256.v1"
)
DATASHEET_JSON_MEDIA_TYPE = "application/json"
DATASHEET_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
DATASHEET_WORKBOOK_SHEETS = (
    "Control",
    "Datasheet",
    "Sources",
    "Assumptions",
    "Calculations",
    "Completeness",
    "Revision",
)
MAX_DATASHEET_EXPORT_BYTES = 8 * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES = 64
MAX_XLSX_UNCOMPRESSED_BYTES = 32 * 1024 * 1024
MAX_XLSX_CELL_CHARACTERS = 32_767
MAX_XLSX_EXACT_INTEGER = (2**53) - 1
_FIXED_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)
_UNSAFE_FORMULA_PREFIXES = frozenset("=+-@")
_ILLEGAL_XML_CONTROL = re.compile("[\x00-\x08\x0b\x0c\x0e-\x1f]")
_EXPECTED_XLSX_MEMBERS = (
    "[Content_Types].xml",
    "_rels/.rels",
    "docProps/app.xml",
    "docProps/core.xml",
    "xl/_rels/workbook.xml.rels",
    "xl/styles.xml",
    "xl/theme/theme1.xml",
    "xl/workbook.xml",
    "xl/worksheets/sheet1.xml",
    "xl/worksheets/sheet2.xml",
    "xl/worksheets/sheet3.xml",
    "xl/worksheets/sheet4.xml",
    "xl/worksheets/sheet5.xml",
    "xl/worksheets/sheet6.xml",
    "xl/worksheets/sheet7.xml",
)


class DatasheetExportError(RuntimeError):
    """Base class for deterministic datasheet export failures."""


class DatasheetExportIntegrityError(DatasheetExportError):
    """Raised when rendered bytes fail a bounded integrity check."""


class DatasheetExportPayload(CalculationModel):
    """Versioned logical JSON document for one immutable revision."""

    schema_id: Literal["engineer4me.datasheet.export.v1"] = DATASHEET_EXPORT_SCHEMA
    schema_version: Literal["1.0.0"] = DATASHEET_EXPORT_VERSION
    revision: DatasheetRevisionRecord
    canonicalization: Literal[
        "engineer4me.datasheet.export.canonical-json.sha256.v1"
    ] = DATASHEET_EXPORT_CANONICALIZATION
    approval_state: Literal["unapproved"] = "unapproved"
    final_design_approval_granted: Literal[False] = False
    standards_conformity_claimed: Literal[False] = False


class DatasheetExportDescriptor(CalculationModel):
    """Checksummed metadata for one exact immutable datasheet revision."""

    schema_id: Literal["engineer4me.datasheet.export.v1"] = DATASHEET_EXPORT_SCHEMA
    schema_version: Literal["1.0.0"] = DATASHEET_EXPORT_VERSION
    canonicalization: Literal[
        "engineer4me.datasheet.export.canonical-json.sha256.v1"
    ] = DATASHEET_EXPORT_CANONICALIZATION
    datasheet_id: UUID
    datasheet_revision_id: UUID
    design_case_id: UUID
    design_revision_id: UUID
    design_revision_number: int = Field(ge=1, le=1_000_000)
    design_revision_fingerprint: FingerprintText
    datasheet_revision_number: int = Field(ge=1, le=100)
    datasheet_revision_fingerprint: FingerprintText
    template_id: Identifier
    template_version: VersionText
    template_fingerprint: FingerprintText
    content_fingerprint: FingerprintText
    completeness_fingerprint: FingerprintText
    json_filename: str = Field(min_length=1, max_length=160)
    json_media_type: Literal["application/json"] = DATASHEET_JSON_MEDIA_TYPE
    json_sha256: FingerprintText
    json_size_bytes: int = Field(ge=1, le=MAX_DATASHEET_EXPORT_BYTES)
    workbook_filename: str = Field(min_length=1, max_length=160)
    workbook_media_type: Literal[
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ] = DATASHEET_XLSX_MEDIA_TYPE
    workbook_sha256: FingerprintText
    workbook_size_bytes: int = Field(ge=1, le=MAX_DATASHEET_EXPORT_BYTES)
    workbook_sheets: tuple[str, ...] = DATASHEET_WORKBOOK_SHEETS
    formula_cells_present: Literal[False] = False
    macros_present: Literal[False] = False
    external_links_present: Literal[False] = False
    approval_state: Literal["unapproved"] = "unapproved"
    final_design_approval_granted: Literal[False] = False
    standards_conformity_claimed: Literal[False] = False

    @model_validator(mode="after")
    def validate_controlled_names_and_sheets(self) -> Self:
        base_name = (
            f"engineer4me-datasheet-{self.datasheet_id}-"
            f"r{self.datasheet_revision_number}-"
            f"{self.datasheet_revision_fingerprint[:12]}"
        )
        if self.json_filename != f"{base_name}.json":
            raise ValueError("JSON filename drifted from the exact revision identity")
        if self.workbook_filename != f"{base_name}.xlsx":
            raise ValueError(
                "workbook filename drifted from the exact revision identity"
            )
        if self.workbook_sheets != DATASHEET_WORKBOOK_SHEETS:
            raise ValueError("workbook sheet metadata drifted")
        return self


@dataclass(frozen=True, slots=True)
class DatasheetExportBundle:
    """In-memory bytes plus their immutable public descriptor."""

    revision: DatasheetRevisionRecord
    descriptor: DatasheetExportDescriptor
    json_bytes: bytes
    workbook_bytes: bytes

    def __post_init__(self) -> None:
        trusted = _trusted_revision(self.revision)
        _verify_descriptor_binding(trusted, self.descriptor)
        validate_datasheet_json_artifact(
            trusted,
            self.descriptor,
            self.json_bytes,
        )
        validate_datasheet_workbook_artifact(
            trusted,
            self.descriptor,
            self.workbook_bytes,
        )


def _trusted_revision(revision: DatasheetRevisionRecord) -> DatasheetRevisionRecord:
    try:
        trusted = DatasheetRevisionRecord.model_validate(
            revision.model_dump(mode="python", round_trip=True, warnings="error")
        )
        content = trusted.snapshot.content
        registered = DEFAULT_DATASHEET_TEMPLATE_REGISTRY.resolve(
            content.template_id,
            content.template_version,
        )
        if (
            registered != trusted.snapshot.template
            or registered.template_fingerprint != content.template_fingerprint
        ):
            raise ValueError("embedded template differs from the controlled registry")
        return trusted
    except DatasheetExportIntegrityError:
        raise
    except (DatasheetTemplateRegistryError, ValueError, TypeError) as exc:
        raise DatasheetExportIntegrityError(
            "The datasheet revision is not bound to its exact controlled template."
        ) from exc


def _verify_descriptor_binding(
    revision: DatasheetRevisionRecord,
    descriptor: DatasheetExportDescriptor,
) -> None:
    content = revision.snapshot.content
    completeness = revision.snapshot.completeness
    if (
        descriptor.datasheet_id != revision.datasheet_id
        or descriptor.datasheet_revision_id != revision.revision_id
        or descriptor.datasheet_revision_number != revision.revision_number
        or descriptor.datasheet_revision_fingerprint != revision.revision_fingerprint
        or descriptor.design_case_id != content.design_case_id
        or descriptor.design_revision_id != content.design_revision_id
        or descriptor.design_revision_number != content.design_revision_number
        or descriptor.design_revision_fingerprint != content.design_revision_fingerprint
        or descriptor.template_id != content.template_id
        or descriptor.template_version != content.template_version
        or descriptor.template_fingerprint != content.template_fingerprint
        or descriptor.content_fingerprint != completeness.content_fingerprint
        or descriptor.completeness_fingerprint != completeness.completeness_fingerprint
    ):
        raise DatasheetExportIntegrityError(
            "The export descriptor is not bound to its exact revision."
        )


def validate_datasheet_json_artifact(
    revision: DatasheetRevisionRecord,
    descriptor: DatasheetExportDescriptor,
    value: bytes,
) -> None:
    """Validate canonical JSON without constructing or rendering a workbook."""

    trusted = _trusted_revision(revision)
    _verify_descriptor_binding(trusted, descriptor)
    if (
        len(value) != descriptor.json_size_bytes
        or _sha256(value) != descriptor.json_sha256
    ):
        raise DatasheetExportIntegrityError(
            "The JSON artifact bytes do not match their descriptor."
        )
    try:
        document = json.loads(value)
        canonical = json.dumps(
            document,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=False,
            allow_nan=False,
        ).encode("utf-8")
        payload = DatasheetExportPayload.model_validate(document)
    except Exception as exc:
        raise DatasheetExportIntegrityError(
            "The canonical JSON artifact could not be validated."
        ) from exc
    if canonical != value or payload.revision != trusted:
        raise DatasheetExportIntegrityError(
            "The canonical JSON artifact is not bound to its exact revision."
        )


def validate_datasheet_workbook_artifact(
    revision: DatasheetRevisionRecord,
    descriptor: DatasheetExportDescriptor,
    value: bytes,
) -> None:
    """Validate exact v1 workbook bytes against the controlled renderer."""

    trusted = _trusted_revision(revision)
    _verify_descriptor_binding(trusted, descriptor)
    if (
        len(value) != descriptor.workbook_size_bytes
        or _sha256(value) != descriptor.workbook_sha256
    ):
        raise DatasheetExportIntegrityError(
            "The workbook artifact bytes do not match their descriptor."
        )
    _verify_workbook_semantics(
        value,
        revision=trusted,
        json_sha256=descriptor.json_sha256,
    )
    # Export schema v1 deliberately pins OpenPyXL and compares the complete
    # normalized artifact. A future renderer must use a new export version and
    # retain this v1 path so historical artifacts remain verifiable.
    expected = render_datasheet_workbook(
        trusted,
        json_sha256=descriptor.json_sha256,
    )
    if value != expected:
        raise DatasheetExportIntegrityError(
            "The workbook artifact differs from the exact controlled v1 rendering."
        )


def build_datasheet_export_payload(
    revision: DatasheetRevisionRecord,
) -> DatasheetExportPayload:
    """Build the complete versioned logical JSON payload."""

    return DatasheetExportPayload(revision=_trusted_revision(revision))


def canonical_datasheet_json(revision: DatasheetRevisionRecord) -> bytes:
    """Serialize the complete revision with stable keys and UTF-8 bytes."""

    payload = build_datasheet_export_payload(revision)
    document = payload.model_dump(
        mode="json",
        round_trip=True,
        warnings="error",
    )
    encoded = json.dumps(
        document,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        allow_nan=False,
    ).encode("utf-8")
    _validate_export_size(encoded, label="JSON")
    return encoded


def _sha256(value: bytes) -> str:
    return sha256(value).hexdigest()


def _validate_export_size(value: bytes, *, label: str) -> None:
    if not value or len(value) > MAX_DATASHEET_EXPORT_BYTES:
        raise DatasheetExportIntegrityError(
            f"The {label} export is empty or exceeds the fixed export limit."
        )


def _safe_cell_text(value: object) -> str:
    """Return non-executable, XML-safe, deterministic display text."""

    text = str(value)
    if _ILLEGAL_XML_CONTROL.search(text):
        raise DatasheetExportIntegrityError(
            "Workbook text contains an illegal XML control character."
        )
    if len(text) > MAX_XLSX_CELL_CHARACTERS:
        raise DatasheetExportIntegrityError(
            "Workbook text exceeds the fixed Excel cell limit."
        )
    candidate = text.lstrip(" \t\r\n\ufeff")
    if text.startswith(("\t", "\r", "\n")) or (
        candidate and candidate[0] in _UNSAFE_FORMULA_PREFIXES
    ):
        return "'" + text
    return text


def _set_cell(
    sheet,
    row: int,
    column: int,
    value: object,
    *,
    text_only: bool = False,
) -> Cell:
    cell = sheet.cell(row=row, column=column)
    if value is None:
        cell.value = ""
        cell.data_type = "s"
        cell.number_format = "@"
    elif text_only or isinstance(value, str):
        cell.value = _safe_cell_text(value)
        cell.data_type = "s"
        cell.number_format = "@"
    elif isinstance(value, bool):
        cell.value = value
        cell.data_type = "b"
    elif isinstance(value, int) and abs(value) > MAX_XLSX_EXACT_INTEGER:
        cell.value = str(value)
        cell.data_type = "s"
        cell.number_format = "@"
    elif type(value) in {int, float}:
        cell.value = value
        cell.data_type = "n"
    else:
        cell.value = _safe_cell_text(value)
        cell.data_type = "s"
        cell.number_format = "@"
    return cell


def _write_table(
    sheet,
    *,
    headers: tuple[str, ...],
    rows: tuple[tuple[object, ...], ...],
) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for column, header in enumerate(headers, start=1):
        cell = _set_cell(sheet, 1, column, header, text_only=True)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            cell = _set_cell(sheet, row_index, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    if headers:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
        )
    _size_columns(sheet, len(headers), len(rows) + 1)


def _size_columns(sheet, columns: int, rows: int) -> None:
    for column in range(1, columns + 1):
        width = 10
        for row in range(1, rows + 1):
            value = sheet.cell(row=row, column=column).value
            if value is not None:
                first_line = str(value).splitlines()[0] if str(value) else ""
                width = max(width, min(48, len(first_line) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width


def _joined(values: tuple[object, ...]) -> str:
    return "\n".join(_safe_cell_text(item) for item in values)


def _field_value(value) -> tuple[object, str, str]:
    if value.state is DatasheetFieldState.UNKNOWN:
        return "UNKNOWN", "", value.unknown_reason or "Unknown"
    if value.state is DatasheetFieldState.NOT_APPLICABLE:
        return "NOT APPLICABLE", "", value.unknown_reason or "Not applicable"
    if isinstance(value.value, EngineeringQuantity):
        quantity = value.value
        notes: list[str] = []
        if quantity.uncertainty is not None:
            notes.append(
                f"Uncertainty: {quantity.uncertainty} {quantity.unit}; "
                f"basis: {quantity.uncertainty_basis}"
            )
        if quantity.significant_figures is not None:
            notes.append(
                f"Presentation significant figures: {quantity.significant_figures}"
            )
        if quantity.decimal_places is not None:
            notes.append(f"Presentation decimal places: {quantity.decimal_places}")
        return format_quantity_value(quantity), quantity.unit, "; ".join(notes)
    if type(value.value) in {int, float}:
        return str(value.value), "", ""
    return value.value, "", ""


def _revision_rows(revision: DatasheetRevisionRecord, json_sha256: str):
    content = revision.snapshot.content
    completeness = revision.snapshot.completeness
    return (
        ("Export schema", DATASHEET_EXPORT_SCHEMA),
        ("Export version", DATASHEET_EXPORT_VERSION),
        ("Datasheet ID", revision.datasheet_id),
        ("Datasheet revision", revision.revision_number),
        ("Datasheet revision ID", revision.revision_id),
        ("Datasheet revision fingerprint", revision.revision_fingerprint),
        ("Supersedes revision ID", revision.supersedes_revision_id or "NONE"),
        (
            "Supersedes revision fingerprint",
            revision.supersedes_revision_fingerprint or "NONE",
        ),
        ("Design case ID", content.design_case_id),
        ("Design revision ID", content.design_revision_id),
        ("Design revision number", content.design_revision_number),
        ("Design revision fingerprint", content.design_revision_fingerprint),
        ("Template ID", content.template_id),
        ("Template version", content.template_version),
        ("Template fingerprint", content.template_fingerprint),
        ("Content fingerprint", completeness.content_fingerprint),
        ("Completeness fingerprint", completeness.completeness_fingerprint),
        ("Completeness state", completeness.state.value),
        ("Ready for review", completeness.ready_for_review),
        ("Lifecycle state", content.lifecycle_state.value),
        ("Change reason", revision.change_reason),
        ("Created by", revision.created_by),
        ("Creator identity origin", revision.creator_origin.value),
        ("Created at UTC", revision.created_at.astimezone(UTC).isoformat()),
        ("Canonical JSON SHA-256", json_sha256),
        ("Approval state", revision.approval_state.value),
        ("Final design approval granted", False),
        ("Standards conformity claimed", False),
        (
            "Engineering status",
            "Preliminary decision support; competent-person review required.",
        ),
    )


def _control_rows(revision: DatasheetRevisionRecord, json_sha256: str):
    content = revision.snapshot.content
    completeness = revision.snapshot.completeness
    rows: list[tuple[object, ...]] = [
        ("CONTROL STATUS", "UNAPPROVED - ENGINEERING REVIEW REQUIRED"),
        ("Lifecycle state", content.lifecycle_state.value),
        ("Completeness state", completeness.state.value),
        ("Ready for review", completeness.ready_for_review),
        ("Datasheet ID", revision.datasheet_id),
        ("Datasheet revision ID", revision.revision_id),
        ("Datasheet revision", revision.revision_number),
        ("Datasheet revision fingerprint", revision.revision_fingerprint),
        ("Design case ID", content.design_case_id),
        ("Design revision ID", content.design_revision_id),
        ("Design revision number", content.design_revision_number),
        ("Design revision fingerprint", content.design_revision_fingerprint),
        ("Template ID", content.template_id),
        ("Template version", content.template_version),
        ("Template fingerprint", content.template_fingerprint),
        ("Content fingerprint", completeness.content_fingerprint),
        ("Completeness fingerprint", completeness.completeness_fingerprint),
        ("Canonical JSON SHA-256", json_sha256),
        (
            "Checksum meaning",
            "Integrity checksum only; this is not a digital signature.",
        ),
        (
            "Missing required field IDs",
            _joined(completeness.missing_required_field_ids) or "NONE",
        ),
        (
            "Unknown required field IDs",
            _joined(completeness.unknown_required_field_ids) or "NONE",
        ),
        (
            "Unconfirmed required field IDs",
            _joined(completeness.unconfirmed_required_field_ids) or "NONE",
        ),
        (
            "Unverified calculation field IDs",
            _joined(completeness.unverified_calculation_field_ids) or "NONE",
        ),
        (
            "Unresolved conditional field IDs",
            _joined(completeness.unresolved_conditional_field_ids) or "NONE",
        ),
        (
            "Optional open field IDs",
            _joined(completeness.optional_open_field_ids) or "NONE",
        ),
        (
            "Not-applicable field IDs",
            _joined(completeness.not_applicable_field_ids) or "NONE",
        ),
        (
            "Unresolved assumption IDs",
            _joined(completeness.unresolved_assumption_ids) or "NONE",
        ),
        (
            "Blocking assumption IDs",
            _joined(completeness.blocking_assumption_ids) or "NONE",
        ),
        ("Final design approval granted", False),
        ("Standards conformity claimed", False),
        (
            "Engineering limitation",
            "Preliminary engineering decision support; competent-person review required.",
        ),
    ]
    if content.template_id == PRESSURE_RELIEF_TEMPLATE_ID:
        rows.append(
            (
                "Pressure-relief limitation",
                "Preliminary sizing support only; not an approved relief-system design.",
            )
        )
    return tuple(rows)


def _datasheet_rows(revision: DatasheetRevisionRecord):
    snapshot = revision.snapshot
    values = {item.field_id: item for item in snapshot.content.field_values}
    assessments = {item.field_id: item for item in snapshot.completeness.assessments}
    section_titles = {
        item.section_id: item.title for item in snapshot.template.sections
    }
    rows: list[tuple[object, ...]] = []
    for definition in snapshot.template.fields:
        value = values[definition.field_id]
        assessment = assessments[definition.field_id]
        display, unit, notes = _field_value(value)
        rows.append(
            (
                section_titles[definition.section_id],
                definition.field_id,
                definition.label,
                definition.requirement.value,
                value.state.value,
                display,
                unit,
                definition.preferred_unit or "",
                value.origin.value,
                _joined(value.source_reference_ids),
                _joined(value.assumption_ids),
                _joined(value.calculation_link_ids),
                assessment.disposition.value,
                assessment.blocking,
                definition.safety_critical,
                notes,
            )
        )
    return tuple(rows)


def _completeness_rows(revision: DatasheetRevisionRecord):
    return tuple(
        (
            item.field_id,
            item.requirement.value,
            "UNRESOLVED" if item.required_now is None else item.required_now,
            item.disposition.value,
            item.blocking,
            item.message,
        )
        for item in revision.snapshot.completeness.assessments
    )


def _source_rows(revision: DatasheetRevisionRecord):
    return tuple(
        (
            item.source_id,
            item.origin.value,
            item.description,
            item.source_revision or "UNKNOWN",
            item.location or "UNKNOWN",
            _joined(item.reference_ids),
        )
        for item in revision.snapshot.content.source_references
    )


def _assumption_rows(revision: DatasheetRevisionRecord):
    return tuple(
        (
            item.assumption_id,
            item.statement,
            item.required_verification,
            item.verification_state.value,
            item.safety_critical,
            _joined(item.source_reference_ids),
            _joined(item.verification_evidence_source_ids),
        )
        for item in revision.snapshot.content.assumptions
    )


def _calculation_rows(revision: DatasheetRevisionRecord):
    return tuple(
        (
            item.link_id,
            item.run_id,
            item.calculation_type,
            item.method_id,
            item.method_version,
            item.result_status.value,
            item.design_case_id,
            item.design_revision_id,
            item.design_revision_number,
            item.design_revision_fingerprint,
            item.output.output_id,
            item.output.name,
            (
                format_quantity_value(item.output.quantity)
                if item.output.quantity is not None
                else item.output.categorical_value
            ),
            item.output.quantity.unit if item.output.quantity is not None else "",
            _joined(item.output.source_step_ids),
            _joined(item.output.source_value_ids),
            _joined(item.output.source_reference_ids),
            item.output.description or "UNKNOWN",
            item.repository_provenance_verified,
            item.source_record_embedded,
            item.historical_link_rewritten,
            item.run_fingerprint,
            item.result_fingerprint,
        )
        for item in revision.snapshot.content.calculation_links
    )


def _workbook_tables(revision: DatasheetRevisionRecord, json_sha256: str):
    return (
        (
            "Control",
            ("Control", "Value"),
            _control_rows(revision, json_sha256),
        ),
        (
            "Datasheet",
            (
                "Section",
                "Field ID",
                "Field",
                "Requirement",
                "State",
                "Value",
                "Stored unit",
                "Preferred unit",
                "Origin",
                "Source reference IDs",
                "Assumption IDs",
                "Calculation link IDs",
                "Completeness disposition",
                "Blocking",
                "Safety critical",
                "Notes",
            ),
            _datasheet_rows(revision),
        ),
        (
            "Sources",
            (
                "Source ID",
                "Origin",
                "Description",
                "Source revision",
                "Location",
                "Reference IDs",
            ),
            _source_rows(revision),
        ),
        (
            "Assumptions",
            (
                "Assumption ID",
                "Statement",
                "Required verification",
                "Verification state",
                "Safety critical",
                "Source IDs",
                "Verification evidence source IDs",
            ),
            _assumption_rows(revision),
        ),
        (
            "Calculations",
            (
                "Link ID",
                "Run ID",
                "Calculation type",
                "Method ID",
                "Method version",
                "Result status",
                "Design case ID",
                "Design revision ID",
                "Design revision number",
                "Design revision fingerprint",
                "Output ID",
                "Output name",
                "Output value",
                "Output unit",
                "Output source step IDs",
                "Output source value IDs",
                "Output source reference IDs",
                "Output description",
                "Repository provenance verified",
                "Source record embedded",
                "Historical link rewritten",
                "Run fingerprint",
                "Result fingerprint",
            ),
            _calculation_rows(revision),
        ),
        (
            "Completeness",
            (
                "Field ID",
                "Requirement",
                "Required now",
                "Disposition",
                "Blocking",
                "Message",
            ),
            _completeness_rows(revision),
        ),
        (
            "Revision",
            ("Property", "Value"),
            _revision_rows(revision, json_sha256),
        ),
    )


def _normalise_zip(raw_workbook: bytes, *, core_timestamp: str) -> bytes:
    source = BytesIO(raw_workbook)
    target = BytesIO()
    with (
        ZipFile(source, "r") as original,
        ZipFile(
            target,
            "w",
            compression=ZIP_STORED,
            strict_timestamps=True,
        ) as normalized,
    ):
        names = original.namelist()
        if len(names) > MAX_XLSX_ARCHIVE_ENTRIES or len(names) != len(set(names)):
            raise DatasheetExportIntegrityError(
                "The workbook archive entry contract is invalid."
            )
        total_uncompressed = sum(item.file_size for item in original.infolist())
        if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
            raise DatasheetExportIntegrityError(
                "The workbook archive exceeds its uncompressed byte limit."
            )
        for name in sorted(names):
            parts = name.replace("\\", "/").split("/")
            if name.startswith(("/", "\\")) or ".." in parts:
                raise DatasheetExportIntegrityError(
                    "The workbook archive contains an unsafe member path."
                )
            data = original.read(name)
            if name == "docProps/core.xml":
                data = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    lambda match: (
                        match.group(1) + core_timestamp.encode("ascii") + match.group(2)
                    ),
                    data,
                    count=1,
                )
            info = ZipInfo(filename=name, date_time=_FIXED_ZIP_TIMESTAMP)
            info.compress_type = ZIP_STORED
            info.create_system = 0
            info.external_attr = 0
            info.comment = b""
            info.extra = b""
            normalized.writestr(
                info,
                data,
                compress_type=ZIP_STORED,
            )
    return target.getvalue()


def render_datasheet_workbook(
    revision: DatasheetRevisionRecord,
    *,
    json_sha256: str | None = None,
) -> bytes:
    """Render one safe workbook and prove that it reopens without formulas."""

    trusted = _trusted_revision(revision)
    canonical_json = canonical_datasheet_json(trusted)
    canonical_checksum = _sha256(canonical_json)
    if json_sha256 is not None and json_sha256 != canonical_checksum:
        raise DatasheetExportIntegrityError(
            "The supplied JSON checksum does not match the revision record."
        )

    workbook = Workbook()
    workbook.remove(workbook.active)
    created = trusted.created_at.astimezone(UTC).replace(tzinfo=None)
    workbook.properties.creator = "Engineer4Me"
    workbook.properties.lastModifiedBy = "Engineer4Me"
    workbook.properties.title = _safe_cell_text(trusted.snapshot.content.title)
    workbook.properties.subject = "Controlled engineering datasheet"
    workbook.properties.description = (
        "Preliminary engineering decision support; not final design approval."
    )
    workbook.properties.created = created
    workbook.properties.modified = created
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"

    sheets = {}
    for name, headers, rows in _workbook_tables(trusted, canonical_checksum):
        sheet = workbook.create_sheet(name)
        _write_table(sheet, headers=headers, rows=rows)
        sheets[name] = sheet
    control_sheet = sheets["Control"]
    control_sheet["A2"].fill = PatternFill("solid", fgColor="C00000")
    control_sheet["A2"].font = Font(color="FFFFFF", bold=True)
    control_sheet["B2"].fill = PatternFill("solid", fgColor="C00000")
    control_sheet["B2"].font = Font(color="FFFFFF", bold=True)

    raw = BytesIO()
    workbook.save(raw)
    rendered = _normalise_zip(
        raw.getvalue(),
        core_timestamp=created.isoformat(timespec="seconds") + "Z",
    )
    _validate_export_size(rendered, label="workbook")
    _verify_workbook(rendered)
    return rendered


def _verify_workbook(value: bytes) -> None:
    try:
        with ZipFile(BytesIO(value), "r") as archive:
            names = archive.namelist()
            if len(names) > MAX_XLSX_ARCHIVE_ENTRIES or len(names) != len(set(names)):
                raise DatasheetExportIntegrityError(
                    "The rendered workbook archive entry contract drifted."
                )
            if tuple(names) != _EXPECTED_XLSX_MEMBERS:
                raise DatasheetExportIntegrityError(
                    "The rendered workbook archive member allowlist drifted."
                )
            if names != sorted(names):
                raise DatasheetExportIntegrityError(
                    "The rendered workbook archive ordering drifted."
                )
            entries = archive.infolist()
            if (
                any(item.date_time != _FIXED_ZIP_TIMESTAMP for item in entries)
                or any(item.compress_type != ZIP_STORED for item in entries)
                or any(item.flag_bits & 0x1 for item in entries)
                or sum(item.file_size for item in entries) > MAX_XLSX_UNCOMPRESSED_BYTES
            ):
                raise DatasheetExportIntegrityError(
                    "The rendered workbook archive metadata drifted."
                )
            forbidden_prefixes = (
                "xl/externalLinks/",
                "xl/embeddings/",
                "xl/activeX/",
                "xl/ctrlProps/",
                "xl/charts/",
                "xl/drawings/",
                "xl/media/",
                "xl/pivotCache/",
                "xl/pivotTables/",
                "xl/slicers/",
                "xl/tables/",
            )
            forbidden_names = {
                "xl/connections.xml",
                "xl/vbaProject.bin",
            }
            for name in names:
                if name in forbidden_names or name.startswith(forbidden_prefixes):
                    raise DatasheetExportIntegrityError(
                        "The rendered workbook contains prohibited active content."
                    )
                data = archive.read(name)
                if name.startswith("xl/worksheets/") and re.search(
                    rb"<(?:[A-Za-z0-9_]+:)?f(?:\s|>)",
                    data,
                ):
                    raise DatasheetExportIntegrityError(
                        "The rendered workbook XML contains a formula."
                    )
                if name.startswith("xl/worksheets/") and re.search(
                    rb"<(?:[A-Za-z0-9_]+:)?(?:conditionalFormatting|dataValidations)(?:\s|>)",
                    data,
                ):
                    raise DatasheetExportIntegrityError(
                        "The rendered workbook contains prohibited rule content."
                    )
                if name.endswith(".rels") and b'TargetMode="External"' in data:
                    raise DatasheetExportIntegrityError(
                        "The rendered workbook contains an external relationship."
                    )
    except DatasheetExportIntegrityError:
        raise
    except Exception as exc:
        raise DatasheetExportIntegrityError(
            "The rendered workbook archive could not be verified."
        ) from exc
    try:
        workbook = load_workbook(
            BytesIO(value),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:  # pragma: no cover - defensive library boundary
        raise DatasheetExportIntegrityError(
            "The rendered workbook could not be reopened."
        ) from exc
    try:
        if tuple(workbook.sheetnames) != DATASHEET_WORKBOOK_SHEETS:
            raise DatasheetExportIntegrityError(
                "The rendered workbook sheet contract drifted."
            )
        if getattr(workbook, "_external_links", []):
            raise DatasheetExportIntegrityError(
                "The rendered workbook unexpectedly contains external links."
            )
        if len(workbook.defined_names):
            raise DatasheetExportIntegrityError(
                "The rendered workbook unexpectedly contains defined names."
            )
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                raise DatasheetExportIntegrityError(
                    "The rendered workbook contains a hidden worksheet."
                )
            if (
                sheet.merged_cells.ranges
                or sheet._images
                or sheet._charts
                or sheet.tables
                or len(sheet.conditional_formatting)
                or sheet.data_validations.count
                or any(
                    dimension.hidden
                    or dimension.collapsed
                    or dimension.outlineLevel
                    or (dimension.height is not None and dimension.height <= 1)
                    for dimension in sheet.row_dimensions.values()
                )
                or any(
                    dimension.hidden
                    or dimension.collapsed
                    or dimension.outlineLevel
                    or (dimension.width is not None and dimension.width <= 1)
                    for dimension in sheet.column_dimensions.values()
                )
            ):
                raise DatasheetExportIntegrityError(
                    "The rendered workbook contains a prohibited presentation overlay."
                )
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        raise DatasheetExportIntegrityError(
                            "The rendered workbook contains an executable formula."
                        )
                    if cell.hyperlink is not None:
                        raise DatasheetExportIntegrityError(
                            "The rendered workbook contains a hyperlink."
                        )
    finally:
        workbook.close()


def _expected_workbook_value(value: object, *, text_only: bool = False):
    if value is None:
        return None
    if text_only or isinstance(value, str):
        rendered = _safe_cell_text(value)
        return rendered if rendered else None
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and abs(value) > MAX_XLSX_EXACT_INTEGER:
        return str(value)
    if type(value) in {int, float}:
        return value
    rendered = _safe_cell_text(value)
    return rendered if rendered else None


def _cell_presentation_signature(cell: Cell):
    return (
        copy(cell.font),
        copy(cell.fill),
        copy(cell.border),
        copy(cell.alignment),
        copy(cell.protection),
        cell.number_format,
        cell.quotePrefix,
        cell.pivotButton,
    )


def _controlled_cell_presentation_signatures():
    workbook = Workbook()
    sheet = workbook.active
    header = _set_cell(sheet, 1, 1, "header", text_only=True)
    header.fill = PatternFill("solid", fgColor="1F4E78")
    header.font = Font(color="FFFFFF", bold=True)
    header.alignment = Alignment(vertical="top", wrap_text=True)
    warning = _set_cell(sheet, 2, 1, "warning", text_only=True)
    warning.fill = PatternFill("solid", fgColor="C00000")
    warning.font = Font(color="FFFFFF", bold=True)
    warning.alignment = Alignment(vertical="top", wrap_text=True)
    text_cell = _set_cell(sheet, 3, 1, "text")
    text_cell.alignment = Alignment(vertical="top", wrap_text=True)
    numeric_cell = _set_cell(sheet, 4, 1, 1)
    numeric_cell.alignment = Alignment(vertical="top", wrap_text=True)
    return {
        "header": _cell_presentation_signature(header),
        "warning": _cell_presentation_signature(warning),
        "text": _cell_presentation_signature(text_cell),
        "numeric": _cell_presentation_signature(numeric_cell),
    }


def _verify_workbook_semantics(
    value: bytes,
    *,
    revision: DatasheetRevisionRecord,
    json_sha256: str,
) -> None:
    """Bind every visible workbook cell to the exact immutable revision."""

    _verify_workbook(value)
    try:
        workbook = load_workbook(
            BytesIO(value),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:  # pragma: no cover - already checked above
        raise DatasheetExportIntegrityError(
            "The rendered workbook could not be reopened for semantic validation."
        ) from exc
    try:
        created = revision.created_at.astimezone(UTC).replace(
            tzinfo=None,
            microsecond=0,
        )
        properties = workbook.properties
        if (
            workbook.active.title != "Control"
            or properties.creator != "Engineer4Me"
            or properties.lastModifiedBy != "Engineer4Me"
            or properties.title != _safe_cell_text(revision.snapshot.content.title)
            or properties.subject != "Controlled engineering datasheet"
            or properties.description
            != "Preliminary engineering decision support; not final design approval."
            or properties.created != created
            or properties.modified != created
            or workbook.calculation.calcMode != "manual"
            or workbook.calculation.fullCalcOnLoad is not False
            or workbook.calculation.forceFullCalc is not False
            or len(workbook.views) != 1
            or workbook.views[0].visibility != "visible"
            or workbook.views[0].minimized
            or not workbook.views[0].showHorizontalScroll
            or not workbook.views[0].showVerticalScroll
            or not workbook.views[0].showSheetTabs
            or workbook.views[0].firstSheet != 0
            or workbook.views[0].activeTab != 0
            or workbook.security.lockStructure
            or workbook.security.lockWindows
            or workbook.security.lockRevision
        ):
            raise DatasheetExportIntegrityError(
                "The workbook presentation entry point drifted."
            )
        controlled_styles = _controlled_cell_presentation_signatures()
        for name, headers, rows in _workbook_tables(revision, json_sha256):
            sheet = workbook[name]
            expected_filter = (
                f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
            )
            view = sheet.sheet_view
            sheet_format = sheet.sheet_format
            if (
                sheet.max_row != len(rows) + 1
                or sheet.max_column != len(headers)
                or any(len(row) != len(headers) for row in rows)
                or sheet.freeze_panes != "A2"
                or sheet.auto_filter.ref != expected_filter
                or sheet.auto_filter.filterColumn
                or sheet.auto_filter.sortState is not None
                or sheet.row_dimensions
                or sheet.protection.sheet
                or sheet.sheet_properties.filterMode
                or view.showFormulas is not None
                or view.showGridLines is not None
                or view.showRowColHeaders is not None
                or view.showZeros is not None
                or view.rightToLeft is not None
                or view.zoomScale is not None
                or view.zoomScaleNormal is not None
                or sheet_format.baseColWidth != 8
                or sheet_format.defaultColWidth is not None
                or sheet_format.defaultRowHeight != 15
                or sheet_format.customHeight is not None
                or sheet_format.zeroHeight is not None
                or sheet_format.outlineLevelRow is not None
                or sheet_format.outlineLevelCol is not None
            ):
                raise DatasheetExportIntegrityError(
                    "The workbook semantic table dimensions drifted."
                )
            expected_rows = (headers, *rows)
            for row_index, expected_row in enumerate(expected_rows, start=1):
                for column_index, expected in enumerate(expected_row, start=1):
                    cell = sheet.cell(row=row_index, column=column_index)
                    expected_value = _expected_workbook_value(
                        expected,
                        text_only=row_index == 1,
                    )
                    if cell.value != expected_value:
                        raise DatasheetExportIntegrityError(
                            "The workbook semantic content drifted from its exact revision."
                        )
                    if expected_value is not None:
                        expected_type = (
                            "b"
                            if isinstance(expected_value, bool)
                            else ("n" if type(expected_value) in {int, float} else "s")
                        )
                        if cell.data_type != expected_type:
                            raise DatasheetExportIntegrityError(
                                "The workbook semantic cell type drifted."
                            )
                    style_kind = (
                        "header"
                        if row_index == 1
                        else (
                            "warning"
                            if name == "Control" and row_index == 2
                            else (
                                "numeric"
                                if isinstance(expected_value, bool)
                                or type(expected_value) in {int, float}
                                else "text"
                            )
                        )
                    )
                    if (
                        _cell_presentation_signature(cell)
                        != controlled_styles[style_kind]
                    ):
                        raise DatasheetExportIntegrityError(
                            "The workbook controlled presentation style drifted."
                        )
            expected_columns = {
                get_column_letter(index) for index in range(1, len(headers) + 1)
            }
            if set(sheet.column_dimensions) != expected_columns:
                raise DatasheetExportIntegrityError(
                    "The workbook semantic column layout drifted."
                )
            for column_index in range(1, len(headers) + 1):
                expected_width = 10
                for expected_row in expected_rows:
                    expected_value = _expected_workbook_value(
                        expected_row[column_index - 1],
                        text_only=expected_row is headers,
                    )
                    if expected_value is not None:
                        first_line = str(expected_value).splitlines()[0]
                        expected_width = max(
                            expected_width,
                            min(48, len(first_line) + 2),
                        )
                actual_width = sheet.column_dimensions[
                    get_column_letter(column_index)
                ].width
                if actual_width != expected_width:
                    raise DatasheetExportIntegrityError(
                        "The workbook semantic column layout drifted."
                    )
    finally:
        workbook.close()


def build_datasheet_export_bundle(
    revision: DatasheetRevisionRecord,
) -> DatasheetExportBundle:
    """Build both exports and their exact checksummed descriptor."""

    trusted = _trusted_revision(revision)
    json_bytes = canonical_datasheet_json(trusted)
    json_checksum = _sha256(json_bytes)
    workbook_bytes = render_datasheet_workbook(
        trusted,
        json_sha256=json_checksum,
    )
    workbook_checksum = _sha256(workbook_bytes)
    base_name = (
        f"engineer4me-datasheet-{trusted.datasheet_id}-"
        f"r{trusted.revision_number}-{trusted.revision_fingerprint[:12]}"
    )
    content = trusted.snapshot.content
    completeness = trusted.snapshot.completeness
    descriptor = DatasheetExportDescriptor(
        datasheet_id=trusted.datasheet_id,
        datasheet_revision_id=trusted.revision_id,
        design_case_id=content.design_case_id,
        design_revision_id=content.design_revision_id,
        design_revision_number=content.design_revision_number,
        design_revision_fingerprint=content.design_revision_fingerprint,
        datasheet_revision_number=trusted.revision_number,
        datasheet_revision_fingerprint=trusted.revision_fingerprint,
        template_id=content.template_id,
        template_version=content.template_version,
        template_fingerprint=content.template_fingerprint,
        content_fingerprint=completeness.content_fingerprint,
        completeness_fingerprint=completeness.completeness_fingerprint,
        json_filename=f"{base_name}.json",
        json_sha256=json_checksum,
        json_size_bytes=len(json_bytes),
        workbook_filename=f"{base_name}.xlsx",
        workbook_sha256=workbook_checksum,
        workbook_size_bytes=len(workbook_bytes),
    )
    return DatasheetExportBundle(
        revision=trusted,
        descriptor=descriptor,
        json_bytes=json_bytes,
        workbook_bytes=workbook_bytes,
    )


__all__ = [
    "DATASHEET_EXPORT_CANONICALIZATION",
    "DATASHEET_EXPORT_SCHEMA",
    "DATASHEET_EXPORT_VERSION",
    "DATASHEET_JSON_MEDIA_TYPE",
    "DATASHEET_WORKBOOK_SHEETS",
    "DATASHEET_XLSX_MEDIA_TYPE",
    "MAX_DATASHEET_EXPORT_BYTES",
    "MAX_XLSX_ARCHIVE_ENTRIES",
    "MAX_XLSX_CELL_CHARACTERS",
    "MAX_XLSX_EXACT_INTEGER",
    "MAX_XLSX_UNCOMPRESSED_BYTES",
    "DatasheetExportBundle",
    "DatasheetExportDescriptor",
    "DatasheetExportError",
    "DatasheetExportIntegrityError",
    "DatasheetExportPayload",
    "build_datasheet_export_bundle",
    "build_datasheet_export_payload",
    "canonical_datasheet_json",
    "render_datasheet_workbook",
    "validate_datasheet_json_artifact",
    "validate_datasheet_workbook_artifact",
]
