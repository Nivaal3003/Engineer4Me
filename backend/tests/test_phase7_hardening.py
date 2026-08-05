"""Deterministic Phase 7 numerical, boundary, and recovery hardening."""

from __future__ import annotations

import ast
from concurrent.futures import ThreadPoolExecutor
from datetime import UTC, datetime
from hashlib import sha256
from io import BytesIO
from math import isfinite, pi, sqrt
from pathlib import Path
from random import Random
from uuid import UUID
from zipfile import ZIP_STORED, ZipFile, ZipInfo

import app.engineering.calculations.dp_flow as dp_flow_module
import pytest
from app.engineering.calculations.dp_flow import (
    DP_FLOW_METHOD_VERSION,
    GENERIC_ORIFICE_METHOD_ID,
    DPFlowConvergenceError,
    DPFlowInputError,
    FlowingFluidProperties,
    TraceableCoefficient,
    assess_generic_orifice_applicability,
    calculate_generic_orifice_flow,
    calculate_permanent_pressure_loss,
    solve_orifice_bore_for_mass_flow,
)
from app.engineering.calculations.dp_flow_workflow_models import (
    GenericOrificeFlowRequest,
)
from app.engineering.calculations.general import (
    GeneralCalculationInputError,
    propagate_independent_uncertainty,
)
from app.engineering.calculations.method_models import NumericApplicabilityRange
from app.engineering.calculations.pressure_relief import PressureReliefFlowBasis
from app.engineering.calculations.pressure_relief_required_area import (
    TraceableReliefAreaCoefficients,
)
from app.engineering.calculations.units import DEFAULT_UNIT_REGISTRY, UnknownUnitError
from app.engineering.design.analyzer_models import (
    analyzer_confidence_band,
    fingerprint_analyzer_payload,
)
from app.engineering.design.datasheet_models import (
    DatasheetContent,
    DatasheetCreateCommand,
    DatasheetFieldOrigin,
    DatasheetFieldState,
    DatasheetFieldValue,
    DatasheetSourceReference,
    evaluate_datasheet_condition,
)
from app.engineering.design.datasheet_registry import PRESSURE_TRANSMITTER_TEMPLATE
from app.engineering.design.datasheet_service import DatasheetService
from app.engineering.design.dp_flow_application_models import (
    DPFlowApplicationRequest,
)
from app.engineering.design.dp_flow_application_wizard import (
    assess_dp_flow_application,
)
from app.engineering.design.xlsx_renderer import (
    DatasheetExportBundle,
    DatasheetExportIntegrityError,
    build_datasheet_export_bundle,
)
from app.services.dp_flow_service import DEFAULT_DP_FLOW_SERVICE
from app.services.dp_flow_service import DPFlowInputError as DPFlowServiceInputError
from openpyxl import load_workbook
from pydantic import ValidationError

_BACKEND = Path(__file__).resolve().parents[1]
_GUARDED_PHASE7_MODULES = (
    "app/engineering/calculations/dp_flow.py",
    "app/engineering/calculations/control_valve.py",
    "app/engineering/calculations/control_valve_compressible.py",
    "app/engineering/calculations/control_valve_installed.py",
    "app/engineering/calculations/pressure_relief.py",
    "app/engineering/calculations/pressure_relief_required_area.py",
    "app/engineering/design/level_application_wizard.py",
    "app/engineering/design/dp_flow_application_wizard.py",
    "app/engineering/design/analyzer_assistant.py",
    "app/services/dp_flow_service.py",
    "app/services/control_valve_service.py",
    "app/services/pressure_relief_service.py",
    "app/services/analyzer_application_service.py",
    "app/engineering/knowledge_calculation_adapter.py",
    "app/engineering/product_selection_requirement_adapter.py",
)
_DYNAMIC_BUILTINS = frozenset({"__import__", "compile", "eval", "exec"})
_DATASHEET_ID = UUID("11300000-0000-4000-8000-000000000001")
_DESIGN_CASE_ID = UUID("11300000-0000-4000-8000-000000000002")
_DESIGN_REVISION_ID = UUID("11300000-0000-4000-8000-000000000003")
_DATASHEET_REVISION_ID = UUID("11300000-0000-4000-8000-000000000004")
_CREATED_AT = datetime(2026, 8, 2, 13, 0, 0, 113000, tzinfo=UTC)


def _coefficient(value: float, label: str) -> TraceableCoefficient:
    return TraceableCoefficient(
        value=value,
        source_reference=f"controlled coefficient record {label}",
        applicable_conditions="the exact generated geometry and operating case",
        supplied_by="independent Phase 7 hardening test",
    )


def _fluid(
    *,
    density: float = 998.2,
    viscosity: float = 1.002e-3,
) -> FlowingFluidProperties:
    return FlowingFluidProperties(
        density_kg_m3=density,
        dynamic_viscosity_pa_s=viscosity,
        pressure_absolute_pa=400_000.0,
        temperature_k=293.15,
        phase="liquid",
        property_source_reference="controlled property record PROP-113",
        condition_basis="the exact generated single-phase flowing condition",
    )


def _orifice_values(**updates: object) -> dict[str, object]:
    values: dict[str, object] = {
        "pipe_inside_diameter_m": 0.1,
        "bore_diameter_m": 0.05,
        "differential_pressure_pa": 10_000.0,
        "fluid": _fluid(),
        "discharge_coefficient": _coefficient(0.61, "CD-113"),
        "expansibility_factor": _coefficient(0.98, "EPS-113"),
    }
    values.update(updates)
    return values


def _application_request() -> DPFlowApplicationRequest:
    return DPFlowApplicationRequest(
        assessment_id="DP-HARDENING-113",
        fluid_phase="liquid",
        objective="process_control",
        pipe_inside_diameter_m=0.1,
        minimum_mass_flow_kg_s=1.0,
        normal_mass_flow_kg_s=5.0,
        maximum_mass_flow_kg_s=10.0,
        flowing_density_kg_m3=998.2,
        flowing_viscosity_pa_s=1.002e-3,
        flowing_absolute_pressure_pa=400_000.0,
        flowing_temperature_k=293.15,
        available_upstream_straight_run_d=20.0,
        available_downstream_straight_run_d=8.0,
        maximum_permanent_pressure_loss_pa=25_000.0,
        required_total_uncertainty_percent=2.0,
        dirty_or_solids_bearing="no",
        erosive="no",
        corrosive="no",
        pulsating_flow="no",
        bidirectional_flow="no",
        wet_gas_or_condensing="no",
        full_pipe_confirmed="yes",
        flashing_or_cavitation_risk="no",
        sonic_or_choked_flow_risk="no",
        intrusive_element_allowed="yes",
        hazardous_area="no",
        sour_or_toxic_service="no",
        oxygen_or_high_purity_service="no",
        approved_standard_or_oem_method_available="yes",
        traceable_coefficient_available="yes",
        include_proprietary_variants=True,
    )


def _execution_request() -> GenericOrificeFlowRequest:
    return GenericOrificeFlowRequest(
        operation="generic_orifice",
        method_id=GENERIC_ORIFICE_METHOD_ID,
        method_version=DP_FLOW_METHOD_VERSION,
        **_orifice_values(),
    )


def _datasheet_revision(hostile_text: str):
    source = DatasheetSourceReference(
        source_id="source-hostile-113",
        origin=DatasheetFieldOrigin.USER_SUPPLIED,
        description=hostile_text,
        reference_ids=("record-113",),
    )
    field = DatasheetFieldValue(
        field_id="service_description",
        state=DatasheetFieldState.KNOWN,
        origin=DatasheetFieldOrigin.USER_SUPPLIED,
        value=hostile_text,
        source_reference_ids=(source.source_id,),
    )
    content = DatasheetContent(
        datasheet_id=_DATASHEET_ID,
        design_case_id=_DESIGN_CASE_ID,
        design_revision_id=_DESIGN_REVISION_ID,
        design_revision_number=1,
        design_revision_fingerprint="1" * 64,
        template_id=PRESSURE_TRANSMITTER_TEMPLATE.template_id,
        template_version=PRESSURE_TRANSMITTER_TEMPLATE.template_version,
        template_fingerprint=PRESSURE_TRANSMITTER_TEMPLATE.template_fingerprint,
        title=hostile_text,
        field_values=(field,),
        source_references=(source,),
    )
    history = DatasheetService().create_history(
        DatasheetCreateCommand(
            content=content,
            change_reason=hostile_text,
            created_by="Phase 7 hardening test",
        ),
        revision_id=_DATASHEET_REVISION_ID,
        created_at=_CREATED_AT,
    )
    return history.revisions[0]


@pytest.mark.parametrize(
    ("field_name", "invalid_value"),
    (
        ("density_kg_m3", -1.0),
        ("dynamic_viscosity_pa_s", 0.0),
        ("density_kg_m3", float("nan")),
    ),
)
def test_dp_kernels_revalidate_bypass_constructed_fluid(
    field_name: str,
    invalid_value: float,
) -> None:
    valid = _fluid()
    payload = valid.model_dump(mode="python", round_trip=True)
    payload[field_name] = invalid_value
    bypassed = FlowingFluidProperties.model_construct(**payload)
    values = _orifice_values(fluid=bypassed)

    applicability = assess_generic_orifice_applicability(**values)
    assert applicability.applicable is False
    assert any("fluid" in reason for reason in applicability.blocking_reasons)
    with pytest.raises(DPFlowInputError):
        calculate_generic_orifice_flow(**values)


def test_dp_loss_revalidates_bypass_constructed_coefficient() -> None:
    valid = _coefficient(0.5, "LOSS-113")
    payload = valid.model_dump(mode="python", round_trip=True)

    for update in (
        {"value": -0.5},
        {"source_reference": "   "},
    ):
        bypassed = TraceableCoefficient.model_construct(**(payload | update))
        with pytest.raises(DPFlowInputError):
            calculate_permanent_pressure_loss(
                measured_differential_pressure_pa=10_000.0,
                permanent_loss_ratio=bypassed,
            )


@pytest.mark.parametrize(
    ("field_name", "invalid_value"),
    (
        ("normal_mass_flow_kg_s", float("nan")),
        ("pipe_inside_diameter_m", -1.0),
        ("fluid_phase", "unsupported-phase"),
        ("include_proprietary_variants", "yes"),
    ),
)
def test_dp_wizard_revalidates_bypass_constructed_request(
    field_name: str,
    invalid_value: object,
) -> None:
    valid = _application_request()
    payload = valid.model_dump(mode="python", round_trip=True)
    payload[field_name] = invalid_value
    bypassed = DPFlowApplicationRequest.model_construct(**payload)

    with pytest.raises((TypeError, ValueError)):
        assess_dp_flow_application(bypassed)


@pytest.mark.parametrize("huge_integer", (10**400, -(10**400)))
def test_huge_integers_fail_closed_at_every_reviewed_numeric_boundary(
    huge_integer: int,
) -> None:
    with pytest.raises(GeneralCalculationInputError):
        propagate_independent_uncertainty((huge_integer,), (1.0,))

    with pytest.raises(ValidationError):
        PressureReliefFlowBasis(
            required_relieving_mass_flow_kg_s=huge_integer,
            load_determination_reference="LOAD-113",
            load_determination_basis="Controlled relieving-load calculation basis.",
            supplied_by="Process engineering",
        )

    with pytest.raises(ValidationError):
        TraceableReliefAreaCoefficients(
            coefficient_set_id="COEFFICIENTS-113",
            discharge_coefficient=huge_integer,
            discharge_coefficient_source_reference="CD-SOURCE-113",
            discharge_coefficient_role="capacity_discharge_coefficient",
            combined_correction_factor=1.0,
            combined_correction_factor_source_reference="K-SOURCE-113",
            combined_correction_factor_role="combined_correction_factor",
            standards_basis_reference="STANDARD-BASIS-113",
            applicable_conditions=(
                "The exact relieving fluid, pressure, and installation basis."
            ),
            supplied_by="Pressure-systems engineer",
            all_required_corrections_included=True,
            double_counting_review_completed=True,
        )

    numeric_range = NumericApplicabilityRange(minimum=-1.0, maximum=1.0)
    assert numeric_range.contains(huge_integer) is False


@pytest.mark.parametrize(
    "score",
    (
        float("nan"),
        float("inf"),
        float("-inf"),
        -1.0,
        100.000_001,
        True,
        "50",
        10**400,
    ),
)
def test_analyzer_confidence_rejects_unbounded_or_coercive_scores(
    score: object,
) -> None:
    with pytest.raises((TypeError, ValueError)):
        analyzer_confidence_band(score)  # type: ignore[arg-type]


def test_analyzer_fingerprint_rejects_ambiguous_and_hostile_payloads() -> None:
    cycle: list[object] = []
    cycle.append(cycle)
    too_deep: object = None
    for _ in range(66):
        too_deep = [too_deep]

    invalid_payloads = (
        {1: "integer key"},
        {"1": "text key", 1: "integer key"},
        {"nonfinite": float("nan")},
        {"unbounded": 10**400},
        {"unsupported": object()},
        cycle,
        too_deep,
    )
    for payload in invalid_payloads:
        with pytest.raises(ValueError):
            fingerprint_analyzer_payload(payload)

    assert fingerprint_analyzer_payload({"value": -0.0}) == (
        fingerprint_analyzer_payload({"value": 0.0})
    )


def test_seeded_orifice_vectors_preserve_metamorphic_invariants() -> None:
    generated = Random(0xE4_07_113)

    for index in range(96):
        pipe_diameter = generated.uniform(0.05, 0.5)
        beta = generated.uniform(0.15, 0.75)
        bore = pipe_diameter * beta
        differential_pressure = generated.uniform(1_000.0, 100_000.0)
        density = generated.uniform(0.8, 1_200.0)
        viscosity = generated.uniform(1.0e-5, 0.1)
        discharge = generated.uniform(0.5, 0.95)
        expansibility = generated.uniform(0.8, 1.0)
        fluid = _fluid(density=density, viscosity=viscosity)
        values = {
            "pipe_inside_diameter_m": pipe_diameter,
            "bore_diameter_m": bore,
            "differential_pressure_pa": differential_pressure,
            "fluid": fluid,
            "discharge_coefficient": _coefficient(
                discharge,
                f"CD-{index}",
            ),
            "expansibility_factor": _coefficient(
                expansibility,
                f"EPS-{index}",
            ),
        }
        result = calculate_generic_orifice_flow(**values)
        expected_mass = (
            discharge
            * expansibility
            * pi
            * bore**2
            / 4.0
            * sqrt(2.0 * differential_pressure * density)
            / sqrt(1.0 - beta**4)
        )

        assert result.mass_flow_kg_s == pytest.approx(expected_mass, rel=2e-13)
        assert result.actual_volumetric_flow_m3_s == pytest.approx(
            result.mass_flow_kg_s / density,
            rel=2e-13,
        )
        assert result.pipe_reynolds_number == pytest.approx(
            4.0 * result.mass_flow_kg_s / (pi * pipe_diameter * viscosity),
            rel=2e-13,
        )
        assert all(
            isfinite(value) and value > 0.0
            for value in (
                result.mass_flow_kg_s,
                result.actual_volumetric_flow_m3_s,
                result.velocity_m_s,
                result.pipe_reynolds_number,
            )
        )

        quadruple_dp = calculate_generic_orifice_flow(
            **(values | {"differential_pressure_pa": 4.0 * differential_pressure})
        )
        assert quadruple_dp.mass_flow_kg_s == pytest.approx(
            2.0 * result.mass_flow_kg_s,
            rel=2e-13,
        )


def test_solver_uses_exactly_two_endpoint_evaluations_plus_iteration_budget(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    values = _orifice_values(bore_diameter_m=0.053_271)
    target = calculate_generic_orifice_flow(**values).mass_flow_kg_s
    original = dp_flow_module.calculate_generic_orifice_flow
    calls = 0

    def counted_calculation(**kwargs: object):
        nonlocal calls
        calls += 1
        return original(**kwargs)

    monkeypatch.setattr(
        dp_flow_module,
        "calculate_generic_orifice_flow",
        counted_calculation,
    )
    iteration_budget = 5
    with pytest.raises(DPFlowConvergenceError, match="did not converge"):
        solve_orifice_bore_for_mass_flow(
            target_mass_flow_kg_s=target,
            pipe_inside_diameter_m=0.1,
            differential_pressure_pa=10_000.0,
            fluid=_fluid(),
            discharge_coefficient=_coefficient(0.61, "CD-SOLVER-113"),
            expansibility_factor=_coefficient(0.98, "EPS-SOLVER-113"),
            minimum_bore_diameter_m=0.01,
            maximum_bore_diameter_m=0.09,
            relative_tolerance=1.0e-15,
            maximum_iterations=iteration_budget,
        )
    assert calls == iteration_budget + 2


def test_generated_malformed_unit_symbols_never_resolve() -> None:
    generated = Random(0xE4_07_113)
    malformed = {
        "__import__('os')",
        "kg / s",
        "m**2",
        "Pа",
        "ｍ",
        "Pa\x00",
        "° C",
    }
    definitions = DEFAULT_UNIT_REGISTRY.definitions
    for index in range(96):
        symbol = definitions[index % len(definitions)].symbol
        nonce = generated.randrange(1, 2**32)
        malformed.add(f"{symbol}::phase7-invalid-{index}-{nonce:08x}")

    for symbol in sorted(malformed):
        with pytest.raises(UnknownUnitError):
            DEFAULT_UNIT_REGISTRY.resolve_unit(symbol)


@pytest.mark.parametrize(
    "hostile_text",
    (
        '=HYPERLINK("https://example.invalid","open")',
        "+cmd|' /C calc'!A0",
        "@SUM(1,1)",
        "  =1+1",
    ),
)
def test_concurrent_datasheet_exports_are_byte_exact_and_formula_safe(
    hostile_text: str,
) -> None:
    revision = _datasheet_revision(hostile_text)
    baseline = build_datasheet_export_bundle(revision)

    def export(_: int) -> tuple[bytes, bytes]:
        bundle = build_datasheet_export_bundle(revision)
        return bundle.json_bytes, bundle.workbook_bytes

    with ThreadPoolExecutor(max_workers=8) as executor:
        artifacts = tuple(executor.map(export, range(16)))

    assert set(artifacts) == {(baseline.json_bytes, baseline.workbook_bytes)}
    workbook = load_workbook(
        BytesIO(baseline.workbook_bytes),
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        normalized = hostile_text.strip()
        matching_cells = [
            cell
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and normalized in cell.value
        ]
        assert matching_cells
        assert all(cell.data_type == "s" for cell in matching_cells)
        assert all(cell.value.startswith("'") for cell in matching_cells)
        assert all(cell.number_format == "@" for cell in matching_cells)
        assert all(
            cell.data_type != "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()


def test_parent_path_archive_member_is_rejected_and_export_recovers() -> None:
    revision = _datasheet_revision("Controlled non-formula recovery case")
    baseline = build_datasheet_export_bundle(revision)
    target = BytesIO()
    with (
        ZipFile(BytesIO(baseline.workbook_bytes), "r") as source,
        ZipFile(target, "w", compression=ZIP_STORED) as forged_archive,
    ):
        for info in source.infolist():
            forged_archive.writestr(info, source.read(info.filename))
        parent_member = ZipInfo(
            "../phase7-parent-path.xml",
            date_time=(1980, 1, 1, 0, 0, 0),
        )
        parent_member.compress_type = ZIP_STORED
        forged_archive.writestr(parent_member, b"<forged />")
    forged_workbook = target.getvalue()
    forged_descriptor = baseline.descriptor.model_copy(
        update={
            "workbook_sha256": sha256(forged_workbook).hexdigest(),
            "workbook_size_bytes": len(forged_workbook),
        }
    )

    with pytest.raises(DatasheetExportIntegrityError):
        DatasheetExportBundle(
            revision=baseline.revision,
            descriptor=forged_descriptor,
            json_bytes=baseline.json_bytes,
            workbook_bytes=forged_workbook,
        )

    recovered = build_datasheet_export_bundle(revision)
    assert recovered == baseline


def test_nonconditional_definition_cannot_enter_condition_evaluator() -> None:
    """Production optimization cannot erase the condition precondition."""

    definition = next(
        item for item in PRESSURE_TRANSMITTER_TEMPLATE.fields if item.condition is None
    )

    with pytest.raises(ValueError, match="condition is required"):
        evaluate_datasheet_condition(
            definition=definition,
            all_values={},
            all_definitions={definition.field_id: definition},
        )


def test_phase7_execution_modules_have_no_dynamic_builtins_or_while_loops() -> None:
    violations: list[str] = []
    for relative_path in _GUARDED_PHASE7_MODULES:
        path = _BACKEND / relative_path
        tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
        for node in ast.walk(tree):
            if isinstance(node, ast.While):
                violations.append(f"{relative_path}:{node.lineno}:while")
            elif (
                isinstance(node, ast.Call)
                and isinstance(node.func, ast.Name)
                and node.func.id in _DYNAMIC_BUILTINS
            ):
                violations.append(f"{relative_path}:{node.lineno}:{node.func.id}")

    assert violations == []


def test_shared_dp_service_recovers_after_concurrent_rejected_requests() -> None:
    valid = _execution_request()
    invalid_payload = valid.model_dump(mode="python", round_trip=True)
    invalid_payload["differential_pressure_pa"] = float("nan")
    invalid = GenericOrificeFlowRequest.model_construct(**invalid_payload)
    baseline = DEFAULT_DP_FLOW_SERVICE.execute(valid)

    def reject_then_execute(_: int) -> str:
        try:
            DEFAULT_DP_FLOW_SERVICE.execute(invalid)
        except DPFlowServiceInputError:
            pass
        else:  # pragma: no cover - explicit fail-closed assertion
            raise AssertionError("invalid shared-service request was accepted")
        return DEFAULT_DP_FLOW_SERVICE.execute(valid).trace.attempt_fingerprint

    with ThreadPoolExecutor(max_workers=8) as executor:
        fingerprints = tuple(executor.map(reject_then_execute, range(32)))

    assert set(fingerprints) == {baseline.trace.attempt_fingerprint}
    assert DEFAULT_DP_FLOW_SERVICE.execute(valid) == baseline
