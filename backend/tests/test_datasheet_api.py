"""Step 110 controlled datasheet persistence and download API tests."""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from hashlib import sha256
from io import BytesIO
import json
from uuid import UUID

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import Engine, create_engine, event
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.db.database import Base, get_db
from app.engineering.design.datasheet_models import (
    DatasheetContent,
    DatasheetCreateCommand,
    DatasheetRevisionCreate,
)
from app.engineering.design.datasheet_registry import (
    PRESSURE_TRANSMITTER_TEMPLATE,
)
from app.engineering.design.persistence_models import (
    DesignCaseCreate,
    DesignRevisionPayload,
)
from app.engineering.design.xlsx_renderer import DATASHEET_WORKBOOK_SHEETS
from app.main import APPLICATION_VERSION, app
from app.models.calculation_run import CalculationRun
from app.models.design_case import DesignCase, DesignCaseRevision
from app.models.engineering_datasheet import (
    EngineeringDatasheet,
    EngineeringDatasheetCalculationLink,
    EngineeringDatasheetRevision,
)


DATASHEET_ID = UUID("30000000-0000-4000-8000-000000000003")


@dataclass(frozen=True)
class _Api:
    client: TestClient
    engine: Engine


@pytest.fixture
def api() -> Iterator[_Api]:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def _foreign_keys(dbapi_connection, _connection_record) -> None:
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    Base.metadata.create_all(
        engine,
        tables=(
            DesignCase.__table__,
            DesignCaseRevision.__table__,
            CalculationRun.__table__,
            EngineeringDatasheet.__table__,
            EngineeringDatasheetRevision.__table__,
            EngineeringDatasheetCalculationLink.__table__,
        ),
    )

    def _database_override():
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_db] = _database_override
    try:
        with TestClient(app) as client:
            yield _Api(client=client, engine=engine)
    finally:
        app.dependency_overrides.clear()
        engine.dispose()


def _create_case(api: _Api) -> dict[str, object]:
    command = DesignCaseCreate(
        case_reference="E4M-API-DATASHEET-110",
        case_type="pressure-transmitter",
        payload=DesignRevisionPayload(
            title="Step 110 controlled datasheet design",
            discipline="process-instrumentation",
        ),
        change_reason="Create the design case for datasheet testing.",
        created_by="API test engineer",
    )
    response = api.client.post(
        "/api/v1/designs",
        json=command.model_dump(mode="json", round_trip=True, warnings="error"),
    )
    assert response.status_code == 201, response.text
    return response.json()


def _create_datasheet(
    api: _Api,
    case: dict[str, object],
    *,
    title: str = "PT-110 API datasheet",
) -> dict[str, object]:
    revision = case["revision"]
    assert isinstance(revision, dict)
    template = PRESSURE_TRANSMITTER_TEMPLATE
    command = DatasheetCreateCommand(
        content=DatasheetContent(
            datasheet_id=DATASHEET_ID,
            design_case_id=UUID(str(case["design_case_id"])),
            design_revision_id=UUID(str(revision["revision_id"])),
            design_revision_number=int(revision["revision_number"]),
            design_revision_fingerprint=str(revision["revision_fingerprint"]),
            template_id=template.template_id,
            template_version=template.template_version,
            template_fingerprint=template.template_fingerprint,
            title=title,
        ),
        change_reason=(
            title if title.startswith(("=", "+", "-", "@")) else "Create sheet."
        ),
        created_by=(
            title if title.startswith(("=", "+", "-", "@")) else "API engineer"
        ),
    )
    case_id = case["design_case_id"]
    response = api.client.post(
        f"/api/v1/designs/{case_id}/datasheets",
        json=command.model_dump(mode="json", round_trip=True, warnings="error"),
    )
    assert response.status_code == 201, response.text
    return response.json()


def test_openapi_has_only_nested_append_read_and_exact_export_routes() -> None:
    document = app.openapi()
    paths = {
        "/api/v1/designs/{design_case_id}/datasheets",
        "/api/v1/designs/{design_case_id}/datasheets/{datasheet_id}",
        "/api/v1/designs/{design_case_id}/datasheets/{datasheet_id}/revisions",
        "/api/v1/designs/{design_case_id}/datasheets/{datasheet_id}/revisions/{revision_number}",
        "/api/v1/designs/{design_case_id}/datasheets/{datasheet_id}/revisions/{revision_number}/exports/{export_format}",
    }
    assert paths.issubset(document["paths"])
    for path in paths:
        assert not ({"put", "patch", "delete"} & set(document["paths"][path]))
    assert all("latest" not in path for path in document["paths"])
    assert document["info"]["version"] == APPLICATION_VERSION == "0.10.0"


def test_create_list_get_and_exact_revision(api: _Api) -> None:
    case = _create_case(api)
    created = _create_datasheet(api, case)
    case_id = case["design_case_id"]
    assert created["datasheet_id"] == str(DATASHEET_ID)
    assert created["current_revision"] == 1
    assert (
        created["current"]["revision"]["snapshot"]["completeness"]["state"] == "blocked"
    )
    assert not created["current"]["revision"]["final_design_approval_granted"]
    assert len(created["current"]["export"]["workbook_sha256"]) == 64

    listed = api.client.get(f"/api/v1/designs/{case_id}/datasheets")
    assert listed.status_code == 200
    assert listed.json()["total"] == 1
    assert "snapshot" not in listed.json()["items"][0]

    current = api.client.get(f"/api/v1/designs/{case_id}/datasheets/{DATASHEET_ID}")
    assert current.status_code == 200
    assert current.json() == created

    exact = api.client.get(
        f"/api/v1/designs/{case_id}/datasheets/{DATASHEET_ID}/revisions/1"
    )
    assert exact.status_code == 200
    assert exact.json() == created["current"]

    revisions = api.client.get(
        f"/api/v1/designs/{case_id}/datasheets/{DATASHEET_ID}/revisions"
    )
    assert revisions.status_code == 200
    assert revisions.json()["total"] == 1
    assert "snapshot" not in revisions.json()["items"][0]


def test_json_and_xlsx_downloads_have_exact_bytes_and_security_headers(
    api: _Api,
) -> None:
    case = _create_case(api)
    hostile = '=HYPERLINK("https://example.invalid","open")'
    created = _create_datasheet(api, case, title=hostile)
    case_id = case["design_case_id"]
    base = f"/api/v1/designs/{case_id}/datasheets/{DATASHEET_ID}/revisions/1/exports"
    json_response = api.client.get(f"{base}/json")
    assert json_response.status_code == 200
    assert json_response.headers["content-type"] == ("application/json; charset=utf-8")
    assert json_response.headers["cache-control"] == "private, no-store, max-age=0"
    assert json_response.headers["pragma"] == "no-cache"
    assert json_response.headers["x-content-type-options"] == "nosniff"
    assert json_response.headers["cross-origin-resource-policy"] == "same-origin"
    json_checksum = sha256(json_response.content).hexdigest()
    assert json_response.headers["x-checksum-sha256"] == json_checksum
    assert json_response.headers["etag"] == f'"{json_checksum}"'
    assert str(DATASHEET_ID) in json_response.headers["content-disposition"]
    document = json.loads(json_response.content)
    revision = document["revision"]
    assert revision["snapshot"]["content"]["title"] == hostile
    assert revision["snapshot"]["completeness"]["state"] == "blocked"
    assert not document["final_design_approval_granted"]

    xlsx_response = api.client.get(f"{base}/xlsx")
    assert xlsx_response.status_code == 200
    assert xlsx_response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    xlsx_checksum = sha256(xlsx_response.content).hexdigest()
    assert xlsx_response.headers["x-checksum-sha256"] == xlsx_checksum
    assert xlsx_checksum == created["current"]["export"]["workbook_sha256"]
    workbook = load_workbook(BytesIO(xlsx_response.content), data_only=False)
    try:
        assert tuple(workbook.sheetnames) == DATASHEET_WORKBOOK_SHEETS
        assert all(
            cell.data_type != "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        assert any(
            isinstance(cell.value, str) and cell.value.startswith("'=HYPERLINK")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()

    repeated = api.client.get(f"{base}/xlsx")
    assert repeated.content == xlsx_response.content


def test_revision_append_is_cas_guarded_and_historical_export_is_stable(
    api: _Api,
) -> None:
    case = _create_case(api)
    created = _create_datasheet(api, case)
    case_id = case["design_case_id"]
    original = api.client.get(
        f"/api/v1/designs/{case_id}/datasheets/{DATASHEET_ID}/revisions/1/exports/json"
    ).content
    content = created["current"]["revision"]["snapshot"]["content"]
    content["title"] = "PT-110 API datasheet revision two"
    command = DatasheetRevisionCreate(
        expected_current_revision=1,
        expected_current_fingerprint=created["current_revision_fingerprint"],
        content=DatasheetContent.model_validate(content),
        change_reason="Record revision two.",
        created_by="API engineer",
    ).model_dump(mode="json", round_trip=True, warnings="error")
    path = f"/api/v1/designs/{case_id}/datasheets/{DATASHEET_ID}/revisions"
    revised = api.client.post(path, json=command)
    assert revised.status_code == 201, revised.text
    assert revised.json()["current_revision"] == 2

    stale = api.client.post(path, json=command)
    assert stale.status_code == 409
    assert stale.json()["detail"]["code"] == "datasheet_persistence_conflict"
    historical = api.client.get(f"{path}/1/exports/json")
    assert historical.status_code == 200
    assert historical.content == original


def test_cross_case_missing_invalid_format_and_body_mismatch_fail_closed(
    api: _Api,
) -> None:
    case = _create_case(api)
    _create_datasheet(api, case)
    case_id = case["design_case_id"]
    other_case = "40000000-0000-4000-8000-000000000001"
    hidden = api.client.get(f"/api/v1/designs/{other_case}/datasheets/{DATASHEET_ID}")
    assert hidden.status_code == 404
    assert hidden.json()["detail"]["code"] == "datasheet_not_found"

    unsupported = api.client.get(
        f"/api/v1/designs/{case_id}/datasheets/{DATASHEET_ID}/revisions/1/exports/pdf"
    )
    assert unsupported.status_code == 422
    assert "traceback" not in unsupported.text.lower()

    revision = case["revision"]
    template = PRESSURE_TRANSMITTER_TEMPLATE
    foreign = DatasheetCreateCommand(
        content=DatasheetContent(
            datasheet_id=UUID("50000000-0000-4000-8000-000000000001"),
            design_case_id=UUID(other_case),
            design_revision_id=UUID(str(revision["revision_id"])),
            design_revision_number=1,
            design_revision_fingerprint=str(revision["revision_fingerprint"]),
            template_id=template.template_id,
            template_version=template.template_version,
            template_fingerprint=template.template_fingerprint,
            title="Foreign datasheet",
        ),
        change_reason="Reject the path mismatch.",
        created_by="API engineer",
    )
    mismatch = api.client.post(
        f"/api/v1/designs/{case_id}/datasheets",
        json=foreign.model_dump(mode="json", round_trip=True, warnings="error"),
    )
    assert mismatch.status_code == 422
    assert mismatch.json()["detail"][0]["type"] == "datasheet_command_invalid"


def test_request_limit_and_unknown_fields_are_rejected(api: _Api) -> None:
    case = _create_case(api)
    case_id = case["design_case_id"]
    too_large = api.client.post(
        f"/api/v1/designs/{case_id}/datasheets",
        content=b"{" + b'"padding":"' + b"x" * (1024 * 1024) + b'"}',
        headers={"content-type": "application/json"},
    )
    assert too_large.status_code == 413
    assert too_large.json()["detail"]["code"] == "design_request_too_large"

    revision = case["revision"]
    template = PRESSURE_TRANSMITTER_TEMPLATE
    document = DatasheetCreateCommand(
        content=DatasheetContent(
            datasheet_id=DATASHEET_ID,
            design_case_id=UUID(str(case_id)),
            design_revision_id=UUID(str(revision["revision_id"])),
            design_revision_number=1,
            design_revision_fingerprint=str(revision["revision_fingerprint"]),
            template_id=template.template_id,
            template_version=template.template_version,
            template_fingerprint=template.template_fingerprint,
            title="Unknown-field test",
        ),
        change_reason="Reject unknown fields.",
        created_by="API engineer",
    ).model_dump(mode="json", round_trip=True, warnings="error")
    document["caller_approval"] = True
    response = api.client.post(
        f"/api/v1/designs/{case_id}/datasheets",
        json=document,
    )
    assert response.status_code == 422
    assert response.json()["detail"] == [
        {
            "type": "extra_forbidden",
            "loc": ["body", "extra_field"],
            "msg": "The request value is invalid.",
        }
    ]


def test_unknown_controlled_template_version_is_not_found(api: _Api) -> None:
    case = _create_case(api)
    revision = case["revision"]
    command = DatasheetCreateCommand(
        content=DatasheetContent(
            datasheet_id=DATASHEET_ID,
            design_case_id=UUID(str(case["design_case_id"])),
            design_revision_id=UUID(str(revision["revision_id"])),
            design_revision_number=1,
            design_revision_fingerprint=str(revision["revision_fingerprint"]),
            template_id=PRESSURE_TRANSMITTER_TEMPLATE.template_id,
            template_version="9.9.9",
            template_fingerprint="0" * 64,
            title="Unknown template version",
        ),
        change_reason="Reject the unavailable exact template version.",
        created_by="API engineer",
    )
    response = api.client.post(
        f"/api/v1/designs/{case['design_case_id']}/datasheets",
        json=command.model_dump(mode="json", round_trip=True, warnings="error"),
    )
    assert response.status_code == 404
    assert response.json()["detail"]["code"] == "datasheet_template_not_found"
