"""Step 110 deterministic JSON and formula-safe workbook export tests."""

from __future__ import annotations

from datetime import UTC, datetime
from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path
from uuid import UUID
from zipfile import ZIP_STORED, ZipFile, ZipInfo

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import ValidationError

from app.engineering.design.datasheet_models import (
    DatasheetContent,
    DatasheetCreateCommand,
    DatasheetAssumption,
    DatasheetFieldOrigin,
    DatasheetFieldState,
    DatasheetFieldValue,
    DatasheetSourceReference,
)
from app.engineering.design.datasheet_registry import (
    PRESSURE_RELIEF_TEMPLATE,
    PRESSURE_TRANSMITTER_TEMPLATE,
)
from app.engineering.design.datasheet_service import DatasheetService
from app.engineering.design.xlsx_renderer import (
    DATASHEET_EXPORT_SCHEMA,
    DATASHEET_EXPORT_VERSION,
    DATASHEET_WORKBOOK_SHEETS,
    MAX_XLSX_EXACT_INTEGER,
    DatasheetExportIntegrityError,
    DatasheetExportBundle,
    DatasheetExportDescriptor,
    _normalise_zip,
    build_datasheet_export_bundle,
    canonical_datasheet_json,
    render_datasheet_workbook,
)


DATASHEET_ID = UUID("aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa")
DESIGN_CASE_ID = UUID("bbbbbbbb-bbbb-4bbb-8bbb-bbbbbbbbbbbb")
DESIGN_REVISION_ID = UUID("cccccccc-cccc-4ccc-8ccc-cccccccccccc")
DATASHEET_REVISION_ID = UUID("dddddddd-dddd-4ddd-8ddd-dddddddddddd")
FINGERPRINT = "1" * 64
CREATED_AT = datetime(2026, 8, 2, 12, 30, tzinfo=UTC)


def _revision(
    *,
    hostile_text: str | None = None,
    numeric_value: int | float | None = None,
    trace_collision: bool = False,
    template=PRESSURE_TRANSMITTER_TEMPLATE,
    created_at: datetime = CREATED_AT,
):
    source_references = ()
    field_values: list[DatasheetFieldValue] = []
    title = "PT-110 controlled datasheet"
    assumptions = ()
    if hostile_text is not None or numeric_value is not None:
        source_text = hostile_text or "Controlled numeric evidence"
        source_references = (
            DatasheetSourceReference(
                source_id="source-user",
                origin=DatasheetFieldOrigin.USER_SUPPLIED,
                description=source_text,
                reference_ids=("record-110",),
            ),
        )
    if hostile_text is not None:
        title = hostile_text
        field_values.append(
            DatasheetFieldValue(
                field_id="service_description",
                state=DatasheetFieldState.KNOWN,
                origin=DatasheetFieldOrigin.USER_SUPPLIED,
                value=hostile_text,
                source_reference_ids=("source-user",),
            )
        )
    if numeric_value is not None:
        field_values.append(
            DatasheetFieldValue(
                field_id="required_accuracy_percent",
                state=DatasheetFieldState.KNOWN,
                origin=DatasheetFieldOrigin.USER_SUPPLIED,
                value=numeric_value,
                source_reference_ids=("source-user",),
            )
        )
    if trace_collision:
        source_references = (
            DatasheetSourceReference(
                source_id="shared-trace-id",
                origin=DatasheetFieldOrigin.USER_SUPPLIED,
                description="Controlled source with a colliding namespace ID.",
                reference_ids=("record-110",),
            ),
        )
        assumptions = (
            DatasheetAssumption(
                assumption_id="shared-trace-id",
                statement="A controlled assumption with the same text ID.",
                required_verification="Verify before engineering review.",
                source_reference_ids=("shared-trace-id",),
            ),
        )
        field_values.append(
            DatasheetFieldValue(
                field_id="minimum_process_pressure",
                state=DatasheetFieldState.UNKNOWN,
                origin=DatasheetFieldOrigin.UNKNOWN,
                source_reference_ids=("shared-trace-id",),
                assumption_ids=("shared-trace-id",),
                unknown_reason="The traced pressure remains unknown.",
            )
        )
    content = DatasheetContent(
        datasheet_id=DATASHEET_ID,
        design_case_id=DESIGN_CASE_ID,
        design_revision_id=DESIGN_REVISION_ID,
        design_revision_number=1,
        design_revision_fingerprint=FINGERPRINT,
        template_id=template.template_id,
        template_version=template.template_version,
        template_fingerprint=template.template_fingerprint,
        title=title,
        field_values=tuple(field_values),
        source_references=source_references,
        assumptions=assumptions,
    )
    history = DatasheetService().create_history(
        DatasheetCreateCommand(
            content=content,
            change_reason=(
                hostile_text
                if hostile_text is not None
                else "Create the controlled datasheet export fixture."
            ),
            created_by=(
                hostile_text if hostile_text is not None else "Export test engineer"
            ),
        ),
        revision_id=DATASHEET_REVISION_ID,
        created_at=created_at,
    )
    return history.revisions[0]


def _open(value: bytes):
    return load_workbook(
        BytesIO(value),
        read_only=False,
        data_only=False,
        keep_links=False,
    )


def test_canonical_json_is_complete_deterministic_and_has_no_runtime_bytes() -> None:
    revision = _revision()
    first = canonical_datasheet_json(revision)
    second = canonical_datasheet_json(revision.model_copy(deep=True))
    assert first == second
    assert not first.endswith(b"\n")
    assert not first.startswith(b"\xef\xbb\xbf")
    assert b"NaN" not in first and b"Infinity" not in first
    document = json.loads(first)
    assert document["schema_id"] == DATASHEET_EXPORT_SCHEMA
    assert document["schema_version"] == DATASHEET_EXPORT_VERSION
    assert document["revision"] == revision.model_dump(
        mode="json",
        round_trip=True,
        warnings="error",
    )
    assert document["revision"]["revision_fingerprint"] == (
        revision.revision_fingerprint
    )
    assert not document["revision"]["snapshot"]["content"][
        "final_design_approval_granted"
    ]
    assert not document["final_design_approval_granted"]
    assert not document["standards_conformity_claimed"]


def test_export_bundle_has_exact_checksums_names_and_fail_closed_flags() -> None:
    bundle = build_datasheet_export_bundle(_revision())
    descriptor = bundle.descriptor
    assert descriptor.schema_id == DATASHEET_EXPORT_SCHEMA
    assert descriptor.schema_version == DATASHEET_EXPORT_VERSION
    assert descriptor.json_sha256 == sha256(bundle.json_bytes).hexdigest()
    assert descriptor.workbook_sha256 == sha256(bundle.workbook_bytes).hexdigest()
    assert descriptor.json_size_bytes == len(bundle.json_bytes)
    assert descriptor.workbook_size_bytes == len(bundle.workbook_bytes)
    fingerprint_prefix = bundle.revision.revision_fingerprint[:12]
    assert descriptor.json_filename == (
        f"engineer4me-datasheet-{DATASHEET_ID}-r1-{fingerprint_prefix}.json"
    )
    assert descriptor.workbook_filename == (
        f"engineer4me-datasheet-{DATASHEET_ID}-r1-{fingerprint_prefix}.xlsx"
    )
    assert ".." not in descriptor.json_filename
    assert "/" not in descriptor.workbook_filename
    assert descriptor.workbook_sheets == DATASHEET_WORKBOOK_SHEETS
    assert not descriptor.formula_cells_present
    assert not descriptor.macros_present
    assert not descriptor.external_links_present
    assert descriptor.approval_state == "unapproved"
    assert not descriptor.final_design_approval_granted
    assert not descriptor.standards_conformity_claimed


def test_workbook_bytes_are_deterministic_and_reopen_with_exact_sheets() -> None:
    revision = _revision()
    first = render_datasheet_workbook(revision)
    second = render_datasheet_workbook(revision.model_copy(deep=True))
    assert first == second
    with ZipFile(BytesIO(first), "r") as archive:
        assert archive.namelist() == sorted(archive.namelist())
        assert all(
            item.date_time == (1980, 1, 1, 0, 0, 0) for item in archive.infolist()
        )
        assert all(item.compress_type == ZIP_STORED for item in archive.infolist())
    workbook = _open(first)
    try:
        assert tuple(workbook.sheetnames) == DATASHEET_WORKBOOK_SHEETS
        assert workbook.properties.creator == "Engineer4Me"
        assert workbook.properties.modified == CREATED_AT.replace(tzinfo=None)
        assert not workbook._external_links
        assert workbook.vba_archive is None
        assert all(
            cell.data_type != "f"
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
    finally:
        workbook.close()


def test_bundle_preserves_nonzero_microsecond_revision_identity() -> None:
    created_at = CREATED_AT.replace(microsecond=416366)
    bundle = build_datasheet_export_bundle(_revision(created_at=created_at))
    assert bundle.revision.created_at == created_at
    assert bundle.workbook_bytes


def test_workbook_exposes_revisions_units_and_every_missing_field() -> None:
    revision = _revision()
    workbook = _open(render_datasheet_workbook(revision))
    try:
        revision_values = {
            row[0].value: row[1].value
            for row in workbook["Revision"].iter_rows(min_row=2)
        }
        assert revision_values["Datasheet revision"] == 1
        assert revision_values["Design revision number"] == 1
        assert revision_values["Datasheet revision fingerprint"] == (
            revision.revision_fingerprint
        )
        assert revision_values["Completeness state"] == "blocked"
        assert revision_values["Final design approval granted"] is False
        assert revision_values["Standards conformity claimed"] is False

        rows = list(workbook["Datasheet"].iter_rows(min_row=2, values_only=True))
        assert len(rows) == len(PRESSURE_TRANSMITTER_TEMPLATE.fields)
        by_id = {row[1]: row for row in rows}
        assert by_id["minimum_process_pressure"][5] == "UNKNOWN"
        assert by_id["minimum_process_pressure"][6] is None
        assert by_id["minimum_process_pressure"][7] == "Pa"
        assert by_id["minimum_process_pressure"][15] == (
            "Not supplied for this datasheet revision."
        )
        assert all(row[4] == "unknown" for row in rows)
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "hostile",
    (
        '=HYPERLINK("https://example.invalid","open")',
        "+cmd|' /C calc'!A0",
        "-2+3",
        "@SUM(1,1)",
        "  =1+1",
        "\t=1+1",
        "\ufeff=1+1",
        "\r\n=1+1",
    ),
)
def test_formula_like_text_is_preserved_in_json_but_inert_in_xlsx(
    hostile: str,
) -> None:
    revision = _revision(hostile_text=hostile)
    bundle = build_datasheet_export_bundle(revision)
    document = json.loads(bundle.json_bytes)["revision"]
    normalized_hostile = hostile.strip()
    assert document["change_reason"] == normalized_hostile
    assert document["snapshot"]["content"]["title"] == normalized_hostile
    workbook = _open(bundle.workbook_bytes)
    try:
        cells = [
            cell
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and normalized_hostile in cell.value
        ]
        assert cells
        assert all(cell.data_type == "s" for cell in cells)
        assert all(cell.value.startswith("'") for cell in cells)
        assert all(cell.number_format == "@" for cell in cells)
        assert all(cell.data_type != "f" for cell in cells)
    finally:
        workbook.close()


def test_trace_identifier_namespaces_remain_unambiguous_in_workbook() -> None:
    workbook = _open(render_datasheet_workbook(_revision(trace_collision=True)))
    try:
        sheet = workbook["Datasheet"]
        headers = tuple(cell.value for cell in sheet[1])
        assert headers[9:12] == (
            "Source reference IDs",
            "Assumption IDs",
            "Calculation link IDs",
        )
        row = next(
            item
            for item in sheet.iter_rows(min_row=2, values_only=True)
            if item[1] == "minimum_process_pressure"
        )
        assert row[9] == "shared-trace-id"
        assert row[10] == "shared-trace-id"
        assert row[11] is None
    finally:
        workbook.close()


def test_json_checksum_mismatch_is_rejected() -> None:
    with pytest.raises(DatasheetExportIntegrityError, match="does not match"):
        render_datasheet_workbook(_revision(), json_sha256="0" * 64)


def test_changed_revision_changes_both_export_checksums() -> None:
    first = build_datasheet_export_bundle(_revision())
    second = build_datasheet_export_bundle(_revision(hostile_text="Safe change"))
    assert first.descriptor.json_sha256 != second.descriptor.json_sha256
    assert first.descriptor.workbook_sha256 != second.descriptor.workbook_sha256


@pytest.mark.parametrize(
    "numeric_value",
    (1.0000000000000002, MAX_XLSX_EXACT_INTEGER + 1, 10**100),
)
def test_number_fields_preserve_exact_model_text(numeric_value: int | float) -> None:
    workbook = _open(render_datasheet_workbook(_revision(numeric_value=numeric_value)))
    try:
        rows = workbook["Datasheet"].iter_rows(min_row=2)
        row = next(
            item for item in rows if item[1].value == "required_accuracy_percent"
        )
        assert row[5].value == str(numeric_value)
        assert row[5].data_type == "s"
        assert row[5].number_format == "@"
    finally:
        workbook.close()


def test_pressure_relief_export_discloses_preliminary_sizing_limit() -> None:
    workbook = _open(
        render_datasheet_workbook(_revision(template=PRESSURE_RELIEF_TEMPLATE))
    )
    try:
        controls = {
            row[0].value: row[1].value
            for row in workbook["Control"].iter_rows(min_row=2)
        }
        assert (
            "not an approved relief-system design"
            in controls["Pressure-relief limitation"]
        )
    finally:
        workbook.close()


def test_illegal_xml_control_fails_before_any_artifact_is_returned() -> None:
    with pytest.raises(DatasheetExportIntegrityError, match="illegal XML control"):
        build_datasheet_export_bundle(_revision(hostile_text="unsafe\x0btext"))


@pytest.mark.parametrize(
    ("field_name", "forged_value"),
    (
        ("json_filename", "../forged.json"),
        ("workbook_filename", "safe.xlsx\r\nX-Forged: true"),
        ("workbook_sheets", ("Datasheet",)),
    ),
)
def test_descriptor_rejects_filename_and_sheet_metadata_drift(
    field_name: str,
    forged_value: object,
) -> None:
    descriptor = build_datasheet_export_bundle(_revision()).descriptor
    payload = descriptor.model_dump(mode="python", round_trip=True)
    payload[field_name] = forged_value
    with pytest.raises(ValidationError):
        DatasheetExportDescriptor.model_validate(payload)


def test_bundle_rejects_rehashed_json_bound_to_different_content() -> None:
    bundle = build_datasheet_export_bundle(_revision())
    document = json.loads(bundle.json_bytes)
    document["revision"]["change_reason"] = "Forged artifact content."
    forged_json = json.dumps(
        document,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
    ).encode("utf-8")
    forged_descriptor = bundle.descriptor.model_copy(
        update={
            "json_sha256": sha256(forged_json).hexdigest(),
            "json_size_bytes": len(forged_json),
        }
    )
    with pytest.raises(DatasheetExportIntegrityError, match="canonical JSON"):
        DatasheetExportBundle(
            revision=bundle.revision,
            descriptor=forged_descriptor,
            json_bytes=forged_json,
            workbook_bytes=bundle.workbook_bytes,
        )


def test_bundle_rejects_rehashed_workbook_rule_content() -> None:
    bundle = build_datasheet_export_bundle(_revision())
    target = BytesIO()
    with (
        ZipFile(BytesIO(bundle.workbook_bytes), "r") as source,
        ZipFile(
            target,
            "w",
            compression=ZIP_STORED,
        ) as forged_archive,
    ):
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                data = data.replace(
                    b"</worksheet>",
                    b'<conditionalFormatting sqref="A1"><cfRule type="expression" '
                    b'priority="1"><formula>1=1</formula></cfRule>'
                    b"</conditionalFormatting></worksheet>",
                )
            forged_archive.writestr(info, data)
    forged_workbook = target.getvalue()
    forged_descriptor = bundle.descriptor.model_copy(
        update={
            "workbook_sha256": sha256(forged_workbook).hexdigest(),
            "workbook_size_bytes": len(forged_workbook),
        }
    )
    with pytest.raises(DatasheetExportIntegrityError, match="rule content"):
        DatasheetExportBundle(
            revision=bundle.revision,
            descriptor=forged_descriptor,
            json_bytes=bundle.json_bytes,
            workbook_bytes=forged_workbook,
        )


def test_bundle_rejects_safe_workbook_rendered_from_another_revision() -> None:
    expected = build_datasheet_export_bundle(_revision(numeric_value=1.25))
    different = build_datasheet_export_bundle(_revision(numeric_value=9.75))
    forged_descriptor = expected.descriptor.model_copy(
        update={
            "workbook_sha256": sha256(different.workbook_bytes).hexdigest(),
            "workbook_size_bytes": len(different.workbook_bytes),
        }
    )
    with pytest.raises(DatasheetExportIntegrityError, match="semantic content"):
        DatasheetExportBundle(
            revision=expected.revision,
            descriptor=forged_descriptor,
            json_bytes=expected.json_bytes,
            workbook_bytes=different.workbook_bytes,
        )


def test_bundle_rejects_hidden_controlled_engineering_row() -> None:
    bundle = build_datasheet_export_bundle(_revision())
    workbook = _open(bundle.workbook_bytes)
    raw = BytesIO()
    try:
        workbook["Datasheet"].row_dimensions[2].hidden = True
        workbook.save(raw)
    finally:
        workbook.close()
    forged_workbook = _normalise_zip(
        raw.getvalue(),
        core_timestamp="2026-08-02T12:30:00Z",
    )
    forged_descriptor = bundle.descriptor.model_copy(
        update={
            "workbook_sha256": sha256(forged_workbook).hexdigest(),
            "workbook_size_bytes": len(forged_workbook),
        }
    )
    with pytest.raises(DatasheetExportIntegrityError, match="presentation overlay"):
        DatasheetExportBundle(
            revision=bundle.revision,
            descriptor=forged_descriptor,
            json_bytes=bundle.json_bytes,
            workbook_bytes=forged_workbook,
        )


def test_bundle_rejects_visually_concealed_engineering_value() -> None:
    bundle = build_datasheet_export_bundle(_revision())
    workbook = _open(bundle.workbook_bytes)
    raw = BytesIO()
    try:
        cell = workbook["Datasheet"]["F2"]
        cell.font = Font(color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        workbook.save(raw)
    finally:
        workbook.close()
    forged_workbook = _normalise_zip(
        raw.getvalue(),
        core_timestamp="2026-08-02T12:30:00Z",
    )
    forged_descriptor = bundle.descriptor.model_copy(
        update={
            "workbook_sha256": sha256(forged_workbook).hexdigest(),
            "workbook_size_bytes": len(forged_workbook),
        }
    )
    with pytest.raises(DatasheetExportIntegrityError, match="presentation style"):
        DatasheetExportBundle(
            revision=bundle.revision,
            descriptor=forged_descriptor,
            json_bytes=bundle.json_bytes,
            workbook_bytes=forged_workbook,
        )


def test_bundle_rejects_preapplied_filter_that_can_conceal_rows() -> None:
    bundle = build_datasheet_export_bundle(_revision())
    workbook = _open(bundle.workbook_bytes)
    raw = BytesIO()
    try:
        workbook["Datasheet"].auto_filter.add_filter_column(4, ["known"])
        workbook.save(raw)
    finally:
        workbook.close()
    forged_workbook = _normalise_zip(
        raw.getvalue(),
        core_timestamp="2026-08-02T12:30:00Z",
    )
    forged_descriptor = bundle.descriptor.model_copy(
        update={
            "workbook_sha256": sha256(forged_workbook).hexdigest(),
            "workbook_size_bytes": len(forged_workbook),
        }
    )
    with pytest.raises(DatasheetExportIntegrityError, match="table dimensions"):
        DatasheetExportBundle(
            revision=bundle.revision,
            descriptor=forged_descriptor,
            json_bytes=bundle.json_bytes,
            workbook_bytes=forged_workbook,
        )


def test_bundle_rejects_zero_default_row_height() -> None:
    bundle = build_datasheet_export_bundle(_revision())
    workbook = _open(bundle.workbook_bytes)
    raw = BytesIO()
    try:
        workbook["Datasheet"].sheet_format.zeroHeight = True
        workbook["Datasheet"].sheet_format.defaultRowHeight = 0
        workbook.save(raw)
    finally:
        workbook.close()
    forged_workbook = _normalise_zip(
        raw.getvalue(),
        core_timestamp="2026-08-02T12:30:00Z",
    )
    forged_descriptor = bundle.descriptor.model_copy(
        update={
            "workbook_sha256": sha256(forged_workbook).hexdigest(),
            "workbook_size_bytes": len(forged_workbook),
        }
    )
    with pytest.raises(DatasheetExportIntegrityError, match="table dimensions"):
        DatasheetExportBundle(
            revision=bundle.revision,
            descriptor=forged_descriptor,
            json_bytes=bundle.json_bytes,
            workbook_bytes=forged_workbook,
        )


def test_bundle_rejects_false_approval_in_print_header() -> None:
    bundle = build_datasheet_export_bundle(_revision())
    workbook = _open(bundle.workbook_bytes)
    raw = BytesIO()
    try:
        workbook[
            "Datasheet"
        ].oddHeader.center.text = "APPROVED FINAL DESIGN - FOR CONSTRUCTION"
        workbook.save(raw)
    finally:
        workbook.close()
    forged_workbook = _normalise_zip(
        raw.getvalue(),
        core_timestamp="2026-08-02T12:30:00Z",
    )
    forged_descriptor = bundle.descriptor.model_copy(
        update={
            "workbook_sha256": sha256(forged_workbook).hexdigest(),
            "workbook_size_bytes": len(forged_workbook),
        }
    )
    with pytest.raises(DatasheetExportIntegrityError, match="controlled v1"):
        DatasheetExportBundle(
            revision=bundle.revision,
            descriptor=forged_descriptor,
            json_bytes=bundle.json_bytes,
            workbook_bytes=forged_workbook,
        )


def test_bundle_rejects_unknown_ooxml_archive_member() -> None:
    bundle = build_datasheet_export_bundle(_revision())
    target = BytesIO()
    with (
        ZipFile(BytesIO(bundle.workbook_bytes), "r") as source,
        ZipFile(
            target,
            "w",
            compression=ZIP_STORED,
        ) as forged_archive,
    ):
        for info in source.infolist():
            forged_archive.writestr(info, source.read(info.filename))
        extra = ZipInfo("customUI/forged.xml", date_time=(1980, 1, 1, 0, 0, 0))
        extra.compress_type = ZIP_STORED
        forged_archive.writestr(extra, b"<customUI />")
    forged_workbook = target.getvalue()
    forged_descriptor = bundle.descriptor.model_copy(
        update={
            "workbook_sha256": sha256(forged_workbook).hexdigest(),
            "workbook_size_bytes": len(forged_workbook),
        }
    )
    with pytest.raises(DatasheetExportIntegrityError, match="member allowlist"):
        DatasheetExportBundle(
            revision=bundle.revision,
            descriptor=forged_descriptor,
            json_bytes=bundle.json_bytes,
            workbook_bytes=forged_workbook,
        )


def test_openpyxl_version_is_exactly_pinned_for_deployed_determinism() -> None:
    backend_root = Path(__file__).resolve().parents[1]
    assert '"openpyxl==3.1.5"' in (backend_root / "pyproject.toml").read_text()
    assert '"openpyxl==3.1.5"' in (backend_root / "Dockerfile").read_text()
